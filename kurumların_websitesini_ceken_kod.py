from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import time
import os

# 1. DOSYA YOLUNU AYARLA
dosya_yolu = r"C:\Users\Ahmet\Desktop\data_scrapping\turk_ozel_okul_oncesi_kurumları_listesi.xlsx"  # ← Kendi yolunla değiştir

# 2. FONKSİYONLAR
def chrome_baslat():
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver

def k12_url_bul(driver, okul_adi):
    k12_urls = []
    try:
        driver.get("https://www.google.com")
        time.sleep(1.5)

        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(okul_adi)
        search_box.send_keys(Keys.RETURN)
        time.sleep(2.5)

        links = driver.find_elements(By.CSS_SELECTOR, "a")
        for link in links:
            href = link.get_attribute("href")
            if href and ".k12" in href:
                if href not in k12_urls:
                    k12_urls.append(href)

    except Exception as e:
        print(f"  ⚠️ Hata: {e}")

    return k12_urls

def yeni_excel_olustur(kaynak_dosya):
    klasor = os.path.dirname(kaynak_dosya)
    dosya_adi = os.path.basename(kaynak_dosya).replace(".xlsx", "")
    yeni_dosya = os.path.join(klasor, f"{dosya_adi}_k12_sonuclar.xlsx")

    kaynak_wb = openpyxl.load_workbook(kaynak_dosya)
    kaynak_ws = kaynak_wb.active

    yeni_wb = openpyxl.Workbook()
    yeni_ws = yeni_wb.active
    yeni_ws.title = "K12 Sonuçlar"

    # Başlıkları kaynak dosyadan al + K12 URL sütunu ekle
    basliklar = []
    for col in range(1, 7):
        deger = kaynak_ws.cell(row=1, column=col).value
        basliklar.append(deger)
    basliklar.append("K12 URL'leri")

    # Başlıkları stillendir
    header_fill = PatternFill(start_color="2B5FB8", end_color="2B5FB8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, baslik in enumerate(basliklar, 1):
        cell = yeni_ws.cell(row=1, column=col_idx, value=baslik)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sütun genişlikleri
    yeni_ws.column_dimensions["A"].width = 8
    yeni_ws.column_dimensions["B"].width = 40
    yeni_ws.column_dimensions["C"].width = 25
    yeni_ws.column_dimensions["D"].width = 50
    yeni_ws.column_dimensions["E"].width = 18
    yeni_ws.column_dimensions["F"].width = 15
    yeni_ws.column_dimensions["G"].width = 60

    return kaynak_ws, yeni_ws, yeni_wb, yeni_dosya

def excel_isle(dosya_yolu, baslangic_satir, bitis_satir):
    kaynak_ws, yeni_ws, yeni_wb, yeni_dosya = yeni_excel_olustur(dosya_yolu)

    driver = chrome_baslat()
    print("✅ Chrome başlatıldı\n")

    yeni_satir = 2

    try:
        for satir in range(baslangic_satir, bitis_satir + 1):
            okul_adi = kaynak_ws.cell(row=satir, column=4).value  # D sütunu

            if not okul_adi:
                print(f"Satır {satir}: Boş, atlanıyor...")
                continue

            print(f"🔍 Satır {satir}: {okul_adi}")

            # Kaynak satırdaki tüm verileri kopyala (A-F)
            for col in range(1, 7):
                deger = kaynak_ws.cell(row=satir, column=col).value
                yeni_ws.cell(row=yeni_satir, column=col, value=deger)

            # K12 URL'lerini bul
            urls = k12_url_bul(driver, okul_adi)

            if urls:
                yeni_ws.cell(row=yeni_satir, column=7).value = ", ".join(urls)
                url_cell = yeni_ws.cell(row=yeni_satir, column=7)
                url_cell.font = Font(color="0563C1", underline="single")
                print(f"  ✅ {len(urls)} URL: {urls[0]}")
            else:
                yeni_ws.cell(row=yeni_satir, column=7).value = "Bulunamadı"
                yeni_ws.cell(row=yeni_satir, column=7).font = Font(color="FF0000")
                print(f"  ❌ Bulunamadı")

            # Zebra satır rengi
            if yeni_satir % 2 == 0:
                row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
                for col in range(1, 8):
                    yeni_ws.cell(row=yeni_satir, column=col).fill = row_fill

            yeni_satir += 1

            # Her 5 satırda bir kaydet
            if satir % 5 == 0:
                yeni_wb.save(yeni_dosya)
                print(f"  💾 Kaydedildi (satır {satir})")

            time.sleep(2)

    except KeyboardInterrupt:
        print("\n⛔ Kullanıcı tarafından durduruldu, mevcut veriler kaydediliyor...")

    finally:
        driver.quit()
        yeni_wb.save(yeni_dosya)
        print(f"\n✅ Tamamlandı!")
        print(f"📁 Yeni dosya: {yeni_dosya}")

# 3. ÇALIŞTIR
if __name__ == "__main__":
    baslangic_satir = int(input("Başlangıç satırı (örn: 2): "))
    bitis_satir = int(input("Bitiş satırı (örn: 50): "))
    excel_isle(dosya_yolu, baslangic_satir, bitis_satir)